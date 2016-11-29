#-*- coding:utf-8 -*-
from openpyxl.reader.excel import load_workbook
import time
import re
import json
import os
import sys
from ftplib import FTP

#����svn
print "�Ե�>>>>>>>>>>>>>>> SVN������"
# os.system("svn update D:\TD_Doc\��ֵ")

FILE_CONFIG = 'excelConfig'#��2��û����
# FILE_CONFIG = ['', 'excelConfig001', 'excelConfig002', 'excelConfig001', 'excelConfig001', 'excelConfig001']#��2��û����
# excel�ļ���
excelMaps = ['Task', 'newbie', 'TD_Achieves', 'lang', 'TD_HERO', 'TD_Home', 'TD_Item', 'TD_Monsters', 'TD_SKILL', 'TD_WareHouse', 'WAVES', 'SUB_WAVES', 'Task-test', 'TD_TOWER', 'tollgates', 'TD_Tavern', 'TD_Hire', 'TD_Soldiers', 'endlessWaves', 'TD_Orders', 'TD_Reward', 'TD_Train', 'TD_SpiritInfo','Endless_Sub_Waves','TD_barracks','TD_Activity','wildMonster','TD_explosion','TD_Expedition','pop', 'TD_PVP', 'TD_NPC', 'TD_Vip', 'TD_Technology', 'TD_Competition', 'TD_Dock'];
#���˵�sheet
fitter_keys = ['levelUp', '', 'game_init', 'Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']#heroStrengthenOdd

def usage():
    print "1������������c:\dumpexcel_2.0.py"
    print "2��ѡ����Ҫ������Excel"
    print "3��ѡ���Ƿ��ϴ���������"
    print "4��ѡ���ϴ�����һ��������"

def main():
    #��ʽ��0,1,2|1|dev_branch
    #Ĭ�ϴ�001�����ȡ

    rName = '''
    #########################�°�Ķ�--->1,�ֶ�����excellĿ¼ 2,ָ���ϴ�����#########################
    �롾ѡ���ļ����(3,15,21)|����Ŀ¼(2.0, ArmorGames)|�ϴ�����(branch, cehua, master)|�Ƿ��ϴ�(1)
    �ο���ʽ��4,7,15|ArmorGames|branch|1��Ĭ���ϴ�branch��ֻ��Ҫ��д��ż��ɣ�

    //////////////////////////////////////////////////////////////////////////////////////////////////////////
    / 0   = Task      | 1   = Newbie      | 2   = Achieves     | 3   = lang               | 4  = Hero        /
    / 5   = Home      | 6   = Item        | 7   = Monsters     | 8   = Skill              | 9  = WareHouse   /
    / 10  = Waves     | 11  = Sub_waves   | 12  = Task-test    | 13  = Tower              | 14 = Tollgates   /
    / 15  = Tavern    | 16  = Hire        | 17  = Soldiers     | 18  = EndlessWaves       | 19 = Orders      /
    / 20  = Reward    | 21  = Train       | 22  = SpiritInfo   | 23  = Endless_Sub_Waves  | 24 = Barracks    /
    / 25  = Activity  | 26  = WildMonster | 27  = Explosion    | 28  = Expedition         | 29 = Pop         /
    / 30  = PVP       | 31  = NPC         | 32  = Vip          | 33  = Technology         | 34 = Competition / 
    / 35  = Dock      | 999 = (ȫ���ϴ�)  |                                                                  /
    //////////////////////////////////////////////////////////////////////////////////////////////////////////
    '''
    rawSomething = raw_input(rName)
    rawSomething = rawSomething.split('|')
    if len(rawSomething) < 1:
        print "������Ų���ȷ�����´��룡"
        usage()
        sys.exit()

    #['0,1,2', '2.0', 'dev_branch', 1]
    #����name
    excelNames = rawSomething[0].split(',')
    if '999' in excelNames:
        excelNames = range(0, 35)
        del excelNames[12]

    #ѡ��汾Ŀ¼
    workSpace = '2.0'
    # Ĭ���ϴ�
    isUpload = 1
    # Ĭ���ϴ�����
    baseUploadDev = 'branch'

    if len(rawSomething) >= 2:
        if rawSomething[1]:
            workSpace = rawSomething[1]
    if len(rawSomething) >= 3:
        if rawSomething[2]:
            baseUploadDev = rawSomething[2]
    if len(rawSomething) >= 4:
        if rawSomething[3]:
            isUpload = rawSomething[3]

    #add dev
    baseUploadDev = 'dev_' + baseUploadDev
    # print workSpace, baseUploadDev, isUpload

    #��ȡ�����ļ�
    isUpload = int(isUpload)
    fileConfig = 'c:\\' + FILE_CONFIG + '.json'
    fileConfig = json.load((file(fileConfig)))

    filePath = fileConfig['filePath'] % workSpace
    # print filePath,excelNames;exit();
    
    isUploadFlag = 0

    for getExcelNameId in excelNames:
        getExcelNameId = int(getExcelNameId)
        # print eName
        #��ȡexcel2007�ļ�            
        filePathName = filePath + excelMaps[getExcelNameId] + '.xlsx'
        if os.path.isfile(filePathName) == False:
            filePathName = filePath + 'TD_' + excelMaps[getExcelNameId] + '.xlsx'
            if os.path.isfile(filePathName) == False:
                continue

        readExcel = load_workbook(filePathName)

        #ȡ��һ�ű�
        sheetNames = readExcel.get_sheet_names()
        savePath = fileConfig['savePath']
        for sheetName in sheetNames:
            # ���˲���Ҫ���������ļ���excel��
            if re.match('#', sheetName) or sheetName in fitter_keys:
                continue

            ws = readExcel.get_sheet_by_name(sheetName)
            columnKey = {}
            allData = {}
            keyData = []

            for _row in range(ws.get_highest_row()):
                # ��ÿ�е�һ����Ϊkey
                firstCellVal = ws.cell(row=_row, column=0).value
                # �����µ��ֵ䣬��������
                if _row > 2:
                    allData[firstCellVal] = {}

                for _column in range(ws.get_highest_column()):
                    # ��ȡ���ݼ��̶����ֶ���
                    firstCellKey = ws.cell(row=2, column=_column).value

                    if not firstCellKey:
                        continue
                    if type(firstCellKey) != unicode:
                        firstCellKey = str(firstCellKey)
                    # ��ȡÿ��ÿ�е��ֶ���ֵ
                    cellVal = ws.cell(row=_row, column=_column).value

                    # print cellVal, _row, _column #��ⱨ������

                    # �ڶ��д洢�ֶ���
                    if _row == 2:
                        if not re.search('\*', firstCellKey):
                            if firstCellKey not in keyData:
                                keyData.append(firstCellKey)
                            columnKey[_column] = cellVal

                    # �ӵڶ��п�ʼ
                    if _row > 2:
                        if re.search('\*', firstCellKey):
                            continue

                        try:
                            cellKey = columnKey[_column]
                            if cellVal == None:
                                cellVal = ''
                            allData[firstCellVal][cellKey] = cellVal
                        except:
                            print '=====error======:',_column
                            raise

            f = open(savePath + sheetName + '.php', 'w+')
            f.write(json.dumps(allData))
            f = open(savePath + sheetName + '.txt', 'w+')
            f.write(json.dumps(keyData))
            f.close()

            os.system('php c:\\phpDecode.php ' + sheetName + '.php')

            #�����һ������������������ֱ��ļ������зֱ�
            if getExcelNameId in [6, 11, 10, 14]:
                # print getExcelNameId , "==================================" , isUploadFlag
                os.system('php c:\\cuttingConfig.php ' + sheetName)
                if getExcelNameId in [11, 10, 14] and isUploadFlag == 0:
                    # print baseUploadDev
                    ftpClass('up', 'wavesInfo', 'wavesInfo', baseUploadDev, fileConfig)
                    isUploadFlag = 1

            if isUpload:
                ftpClass('up', sheetName, sheetName, baseUploadDev, fileConfig)

    if isUpload == 0:
        os.system('c:\\copyGit_Svn.py ')

# FTP�ļ�����
def ftpClass(type, baseFile, serverFile, devName, fileConfig):
    host = fileConfig['server']['host']
    user = fileConfig['server']['user']
    pwd = fileConfig['server']['pwd']

    ftp = FTP()
    ftp.connect(host)
    ftp.login(user, pwd)

    # config name ----> see in xml
    #����ȽϿӵ���Unicode��string ��һ���Ȼ����.һ���Ƚ�Unicode ת��Ϊstring
    configPath = fileConfig['configName'].encode("utf-8")
    # print configPath, devName

    basePath = fileConfig['basePath']
    serverPath = fileConfig['serverPath']
    serverPath = serverPath % (devName, configPath)

    basePath = basePath % configPath
    baseFile = basePath + '/' + baseFile
    serverFile = serverPath + '/' + serverFile
    # print baseFile, "||" , serverFile

    if type == 'up':
        if os.path.isfile(baseFile):
            ftpUp(ftp, baseFile, serverFile)
        if os.path.isdir(baseFile):
            fNames = os.listdir(baseFile)

            for fName in fNames:
                if fName != '.svn':
                    ftpUp(ftp, baseFile + '/' + fName, serverFile + '/' + fName, devName, configPath)

    if type == 'down':
        ftpDown(ftp, baseFile, serverFile)

    ftp.close()


# FTP�ļ��ϴ�
def ftpUp(ftp, baseFile, serverFile, devName, configPath):
    f = open(baseFile, 'r')
    # print baseFile, serverFile
    print "�ɹ��ϴ���-->" +devName +"("+configPath+ ")��"+ ftp.storlines('STOR %s'% ('/' + serverFile), f)
    f.close()


# FTP�ļ�����
def ftpDown(ftp, baseFile, serverFile, devName):
    f = open(baseFile, 'wb')
    print "�ɹ��ϴ�["+baseFile+"]����" + ftp.retrbinary('RETR %s'% ('/' + serverFile), f.write)
    f.close()

if __name__ == '__main__':
    #��ʼʱ��
    startTime = time.time()

    opts = sys.argv[1:]
    for op in opts:
        print op
        if op:
            usage()
            sys.exit()

    main()
    print "\n��ʱ:",str("%.2f" % (time.time() - startTime)) + "s."