#-*- coding:utf-8 -*-
from openpyxl.reader.excel import load_workbook
import time
import re
import json
import os
import sys
import paramiko
from ftplib import FTP

#����svn
print "�Ե�>>>>>>>>>>>>>>> SVN������"
os.system("svn update D:\TD_Doc\TD2����Ŀ¼\�����ĵ�")

# excel�ļ���
excelMaps = [
    'Task_VA', 'newbie_VA', 'TD_Achieves', 'lang_VA', 'HERO_VA', 'Home_VA', 'Item_VA', 'Monsters_VA',
    'SKILL_VA', 'TD_WareHouse', 'WAVES_VA', 'SUB_WAVES_VA', 'Task-test', 'TOWER_VA', 'tollgates_VA',
    'Tavern_VA', 'TD_Hire', 'Soldiers_VA', 'endlessWaves', 'Orders_VA', 'Reward_VA', 'Train_VA',
    'TD_SpiritInfo', 'Endless_Sub_Waves', 'barracks_VA', 'Activity_VA', 'wildMonster_VA', 'TD_explosion',
    'TD_Expedition', 'pop', 'PVP_VA', 'NPC_VA', 'Vip_VA', 'TD_Technology', 'TD_Competition', 'Dock_VA',
    'gameSetconfig', 'dropInfo_VA','Port_VA','Dragonester_VA','Totem_VA','HerrCountry_VA','Ornament_VA',
	'Union_VA'
]

# excelMaps = ['Task', 'newbie', 'TD_Achieves', 'lang', 'TD_HERO', 'TD_Home', 'TD_Item', 'TD_Monsters', 'TD_SKILL', 'TD_WareHouse', 'WAVES', 'SUB_WAVES', 'Task-test', 'TD_TOWER', 'tollgates', 'TD_Tavern', 'TD_Hire', 'TD_Soldiers', 'endlessWaves', 'TD_Orders', 'TD_Reward', 'TD_Train', 'TD_SpiritInfo','Endless_Sub_Waves','TD_barracks','TD_Activity','wildMonster','TD_explosion','TD_Expedition','pop', 'TD_PVP', 'TD_NPC', 'TD_Vip', 'TD_Technology', 'TD_Competition', 'TD_Dock'];

#���˵�sheet
fitter_keys = ['levelUp', '', 'game_init', 'Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']
#������ļ������֣�ͨ�����ú�ƽ̨�������ã�
common_keys = ['worldBoss', 'gameSetconfig', 'competitionProgress', 'orders', 'unlock']
#���������ļ�
special_keys = ['gameSetconfig']

def usage():
    print "1������������c:\dumpexcel_2.0.py"
    print "2��ѡ����Ҫ������Excel"
    print "3��ѡ���Ƿ��ϴ���������"
    print "4��ѡ���ϴ�����һ��������"


def main():
    #��ʽ��0,1,2|1|dev_branch
    #Ĭ�ϴ�001�����ȡ?

    rName = '''
    #########################�°�Ķ�--->1,�ֶ�����excellĿ¼ 2,ָ���ϴ�����#########################
    �롾ѡ���ļ����(3,15,21)|�ϴ��汾(2.0, ArmorGames)|�ϴ�����(branch, cehua, master)|�Ƿ��ϴ�(1)
    �ο���ʽ:4,7|1|2.0|branch(Ĭ��branch)     ��TD1 - kot��
    �ο���ʽ:3,2|2|EN|release(Ĭ��release)    ��TD2 - vikingage��

    ////////////////////////////////////////////////////////////////////////////////////////////////////////////
    / 0   = Task      | 1   = Newbie        | 2   = Achieves     | 3   = lang               | 4  = Hero        /
    / 5   = Home      | 6   = Item          | 7   = Monsters     | 8   = Skill              | 9  = WareHouse   /
    / 10  = Waves     | 11  = Sub_waves     | 12  = Task-test    | 13  = Tower              | 14 = Tollgates   /
    / 15  = Tavern    | 16  = Hire          | 17  = Soldiers     | 18  = EndlessWaves       | 19 = Orders      /
    / 20  = Reward    | 21  = Train         | 22  = SpiritInfo   | 23  = Endless_Sub_Waves  | 24 = Barracks    /
    / 25  = Activity  | 26  = WildMonster   | 27  = Explosion    | 28  = Expedition         | 29 = Pop         /
    / 30  = PVP       | 31  = NPC           | 32  = Vip          | 33  = Technology         | 34 = Competition /
    / 35  = Dock      | 36  = gameSetconfig | 37  = dropInfos    | 38  = Port               | 39 = Dragonester /
    / 40  = Totem     | 41  = HerrCountry   | 42  = Ornament     | 43  = Union_VA           | 999 = (ȫ���ϴ�) /
    ////////////////////////////////////////////////////////////////////////////////////////////////////////////
    '''
    rawSomething = raw_input(rName)
    rawSomething = rawSomething.split('|')
    if len(rawSomething) < 1:
        print "������Ų���ȷ�����´��룡"
        usage()
        sys.exit()

    #start
    startTime = time.time()

    #['0,1,2', '2.0', 'dev_branch', 1]
    #����name
    excelNames = rawSomething[0].split(',')
    if '999' in excelNames:
        excelNames = range(0, 36)
        del excelNames[12]

    # Ĭ���ϴ�
    isUpload = 1
    #ѡ��汾Ŀ¼
    workSpace = 'EN'
    # Ĭ���ϴ�����
    baseUploadDev = 'branch'
    # �ϴ��汾
    cur_version = 2

    if len(rawSomething) >= 2:
        if rawSomething[1]:
            cur_version = int(rawSomething[1])
    if len(rawSomething) >= 3:
        if rawSomething[2]:
            workSpace = rawSomething[2]
    if len(rawSomething) >= 4:
        if rawSomething[3]:
            baseUploadDev = rawSomething[3]

    if cur_version == 1:
        baseUploadDev = 'dev_' + baseUploadDev
        fileConfig = json.load((file('c:\excelConfig.json')))
    else:
        if baseUploadDev == 'branch':#Ĭ��
            baseUploadDev = 'release'
        fileConfig = json.load((file('c:\excelConfig-vikingage.json')))

    filePath = fileConfig['filePath'] % workSpace
    # print type(cur_version)
    # print filePath

    #����ȫ������excel
    for getExcelNameId in excelNames:
        getExcelNameId = int(getExcelNameId)

        #�����ж�����
        if excelMaps[getExcelNameId] not in special_keys:
            isUploadFlag = 0

            #��ȡexcel2007�ļ�
            filePathName = filePath + excelMaps[getExcelNameId] + '.xlsx'
            if os.path.isfile(filePathName) == False:
                filePathName = filePath + 'TD_' + excelMaps[getExcelNameId] + '.xlsx'
                if os.path.isfile(filePathName) == False:
                    continue

            readExcel = load_workbook(filePathName)

            #ȡ��һ�ű�
            sheetNames = readExcel.get_sheet_names()
            savePath = fileConfig['savePath']
            for sheetName in sheetNames:
                # ���˲���Ҫ���������ļ���excel��
                if re.match('#', sheetName) or sheetName in fitter_keys:
                    continue

                ws = readExcel.get_sheet_by_name(sheetName)
                columnKey = {}
                allData = {}
                keyData = []

                for _row in range(ws.get_highest_row()):
                    # ��ÿ�е�һ����Ϊkey
                    firstCellVal = ws.cell(row=_row, column=0).value
                    # �����µ��ֵ䣬��������
                    if _row > 2:
                        allData[firstCellVal] = {}

                    for _column in range(ws.get_highest_column()):
                        # ��ȡ���ݼ��̶����ֶ���
                        firstCellKey = ws.cell(row=2, column=_column).value

                        if not firstCellKey:
                            continue
                        if type(firstCellKey) != unicode:
                            firstCellKey = str(firstCellKey)
                            # ��ȡÿ��ÿ�е��ֶ���ֵ
                        cellVal = ws.cell(row=_row, column=_column).value

                        # print cellVal, _row, _column #��ⱨ������

                        # �ڶ��д洢�ֶ���
                        if _row == 2:
                            if not re.search('\*', firstCellKey):
                                if firstCellKey not in keyData:
                                    keyData.append(firstCellKey)
                                columnKey[_column] = cellVal

                        # �ӵڶ��п�ʼ
                        if _row > 2:
                            if re.search('\*', firstCellKey):
                                continue

                            try:
                                cellKey = columnKey[_column]
                                if cellVal == None:
                                    cellVal = ''
                                allData[firstCellVal][cellKey] = cellVal
                            except:
                                print '=====error======:', _column
                                raise

                f = open(savePath + sheetName + '.php', 'w+')
                f.write(json.dumps(allData))
                f = open(savePath + sheetName + '.txt', 'w+')
                f.write(json.dumps(keyData))
                f.close()

                os.system('php c:\\phpDecode.php ' + sheetName + '.php ' + str(cur_version))

                #�����һ������������������ֱ��ļ������зֱ�
                if getExcelNameId in [6, 11, 10, 14]:
                    # print getExcelNameId , "==================================" , isUploadFlag
                    os.system('php c:\\cuttingConfig.php ' + sheetName + ' ' + str(cur_version))
                    if getExcelNameId in [11, 10, 14] and isUploadFlag == 0:
                        # print baseUploadDev
                        ftpClass('up', 'wavesInfo', 'wavesInfo', baseUploadDev, fileConfig, cur_version)
                        isUploadFlag = 1

                #ftp
                if isUpload:
                    ftpClass('up', sheetName, sheetName, baseUploadDev, fileConfig, cur_version)

        else:
            filePathName = filePath + excelMaps[getExcelNameId] + '.xlsx'
            # print filePathName
            readExcel = load_workbook(filePathName)

            sheetNames = readExcel.get_sheet_names()
            savePath = fileConfig['savePath']
            for sheetName in sheetNames:
                if re.match('#', sheetName) or sheetName in fitter_keys:
                    continue

                ws = readExcel.get_sheet_by_name(sheetName)
                allData = {}
                keyData = []

                for _row in range(ws.get_highest_row()):
                    if _row > 0:
                        firstCellVal = ws.cell(row=_row, column=0).value
                        allData[firstCellVal] = {}

                        allData[firstCellVal] = [ws.cell(row=_row, column=1).value, ws.cell(row=_row, column=2).value]
                        keyData.append(firstCellVal)

                f = open(savePath + sheetName + '.php', 'w+')
                f.write(json.dumps(allData))
                f = open(savePath + sheetName + '.txt', 'w+')
                f.write(json.dumps(keyData))
                f.close()

                os.system('php c:\\phpDecode-gameSet.php ' + sheetName + '.php ' + str(cur_version))

                #ftp
                if isUpload:
                    ftpClass('up', sheetName, sheetName, baseUploadDev, fileConfig, cur_version)

                    # if isUpload == 0:
                    #     os.system('c:\\copyGit_Svn.py ')

    print "\n��ʱ:", str("%.2f" % (time.time() - startTime)) + "s."

# FTP�ļ�����
def ftpClass(type, baseFile, serverFile, devName, fileConfig, cur_version):
    host = fileConfig['server']['host']
    user = fileConfig['server']['user']
    pwd = fileConfig['server']['pwd']

    # config name ----> see in xml
    #����ȽϿӵ���Unicode��string ��һ���Ȼ����.һ���Ƚ�Unicode ת��Ϊstring
    configPath = fileConfig['configName'].encode("utf-8")

    if cur_version == 1:
        ftp = FTP()
        ftp.connect(host)
        ftp.login(user, pwd)
    else:
        # ssh_connect(host, 22, user, pwd)
        # client = paramiko.Transport((host, 22))
        # client.connect(username=user, password=pwd)
        # ftp = paramiko.SFTPClient.from_transport(client)
        ftp = FTP()
        ftp.connect(host)
        ftp.login(user, pwd)

        # ���⴦��viking-age
        if baseFile not in common_keys:
            configPath = 'commonConfig'
            # print configPath, baseFile

    # ����д��Ϊ1����Ϊǿ������ά�޸�Ϊftp����ʹ��sftp (ȥ��ע�Ϳɼ���sftp)
    cur_version = 1

    basePath = fileConfig['basePath']
    serverPath = fileConfig['serverPath']
    serverPath = serverPath % (devName, configPath)

    basePath = basePath % configPath
    baseFile = basePath + '/' + baseFile
    serverFile = serverPath + '/' + serverFile
    # print baseFile, os.path.isdir(baseFile)

    if type == 'up':
        if os.path.isfile(baseFile):
            ftpUp(ftp, baseFile, serverFile, configPath, cur_version)
        if os.path.isdir(baseFile):
            fNames = os.listdir(baseFile)

            for fName in fNames:
                if fName != '.svn':
                    ftpUp(ftp, baseFile + '/' + fName, serverFile + '/' + fName, devName, configPath, cur_version)

    if type == 'down':
        ftpDown(ftp, baseFile, serverFile, cur_version)

    if cur_version == 1:
        ftp.close()
    else:
        client.close()

# FTP�ļ��ϴ�
def ftpUp(ftp, baseFile, serverFile, devName, configPath, cur_version):
    # print baseFile, serverFile
    if cur_version == 1:
        f = open(baseFile, 'r')
        result = ftp.storlines('STOR %s' % ('/' + serverFile), f)
        f.close()
    else:
        result = ftp.put(baseFile, serverFile)

    # print type(result)
    print "�ɹ��ϴ���-[" + baseFile.encode("utf-8")[-9:] + "]->" + devName + "(" + configPath + ")"

# FTP�ļ�����
def ftpDown(ftp, baseFile, serverFile, cur_version):
    if cur_version == 1:
        f = open(baseFile, 'wb')
        result = ftp.retrbinary('RETR %s' % ('/' + serverFile), f.write)
        f.close()
    else:
        result = ftp.get(baseFile, serverFile)

    print "�ɹ�����-[" + baseFile.encode("utf-8")[-9:] + "]->" + devName + "(" + configPath + ")"


def ssh_connect(server_ip, server_port, server_user, server_passwd):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(server_ip, server_port, server_user, server_passwd)
    return ssh


def ssh_disconnect(client):
    client.close()


def win_to_linux(sftp, localpath, remotepath):
    '''
    windows��linux�������ϴ��ļ�.
    localpath  Ϊ�����ļ��ľ���·�����磺D:\test.py
    remotepath Ϊ�������˴���ϴ��ļ��ľ���·��,������һ��Ŀ¼���磺/tmp/my_file.txt
    '''
    sftp.put(localpath, remotepath)


def linux_to_win(sftp, localpath, remotepath):
    '''
    ��linux�����������ļ�������
    localpath  Ϊ�����ļ��ľ���·�����磺D:\test.py
    remotepath Ϊ�������˴���ϴ��ļ��ľ���·��,������һ��Ŀ¼���磺/tmp/my_file.txt
    '''
    sftp.get(remotepath, localpath)


if __name__ == '__main__':
    opts = sys.argv[1:]
    for op in opts:
        print op
        if op:
            usage()
            sys.exit()

    main()